from decimal import Decimal

import pytest
from django.contrib.auth import get_user_model
from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from openpyxl import Workbook

from walkie_dokie.contract_intelligence.boq import (
    import_boq_spec,
    trust_boq_item_records,
    trust_boq_summary_records,
)
from walkie_dokie.contract_intelligence.ingestion import run_parser_for_source
from walkie_dokie.contract_intelligence.models import (
    BoqImportSpec,
    BoqItemRecord,
    BoqSummaryRecord,
    Document,
    DocumentVersion,
    KnowledgeProject,
    SourceFile,
)


_SHEETS = (
    "招标清单列表",
    "1-封皮",
    "2-编制说明",
    "3-汇总表",
    "通用工程类-实体工程量清单-东塔",
    "通用工程类-实体工程量清单-西塔",
    "开办费清单",
    "附表01-安全文明施工费",
    "附表02-垂直运输机械",
    "附表03-脚手架",
    "附表04-临时使用永久电梯",
    "附表05-试验与检验",
    "附表06-其他",
    "照管费清单",
    "计日工工程量清单",
    "暂列清单",
    "通用工程类-实体工程量清单-东塔-综合单价分析表",
    "通用工程类-实体工程量清单-西塔-综合单价分析表",
)

_ENTITY_HEADERS = (
    "itemId",
    "parentId",
    "leafFlag",
    "itemSortCode",
    "唯一标识",
    "清单标识",
    "数据来源标识",
    "自定义项导入",
    "项目编号",
    "项目名称",
    "项目特征描述",
    "类别",
    "计量单位",
    "工程数量",
    "不含增值税综合单价（元）",
    "不含增值税汇总合价（元）",
    "备注",
)

_GENERIC_HEADERS = (
    "id",
    "itemId",
    "parentId",
    "leafFlag",
    "项目编码",
    "项目名称",
    "项目特征描述",
    "类别",
    "计量单位",
    "工程量",
    "不含税单价（元）",
    None,
    None,
    "不含税合价（元）",
    "说明",
)


def _append_row(sheet, row_number, values):
    for column_number, value in enumerate(values, 1):
        sheet.cell(row=row_number, column=column_number, value=value)


def _boq_xlsx_bytes(tmp_path) -> bytes:
    path = tmp_path / "boq.xlsx"
    workbook = Workbook()
    workbook.active.title = _SHEETS[0]
    for sheet_name in _SHEETS[1:]:
        workbook.create_sheet(sheet_name)

    manifest = workbook["招标清单列表"]
    _append_row(manifest, 1, ["contractType", "sheet名称"])
    for row_number, sheet_name in enumerate(_SHEETS, 2):
        manifest.cell(row=row_number, column=2, value=sheet_name)

    cover = workbook["1-封皮"]
    cover["A3"] = "投 标 总 价"
    cover["B5"] = "建设单位："
    cover["D5"] = "测试甲方有限公司"
    cover["B6"] = "工程名称："
    cover["D6"] = "测试泛光照明工程"
    cover["B12"] = "增值税税率："
    cover["D12"] = "0.09"
    cover["B14"] = "增值税合价："
    cover["D14"] = "37.8"
    cover["B18"] = "合同含增值税总价B："
    cover["D18"] = "457.8"

    workbook["2-编制说明"]["A1"] = "编制说明"

    summary = workbook["3-汇总表"]
    _append_row(summary, 3, ["项目序号", "项目", "不含增值税合价", "备注"])
    summary_rows = (
        ("一", "开办费清单", "10"),
        ("二", "实体工程量清单", "400"),
        ("1", "东塔", "200"),
        ("2", "西塔", "200"),
        ("三", "照管费清单", None),
        ("四", "计日工工程量清单", "10"),
        ("五", "暂列清单", None),
        (None, "不含增值税总计", "420"),
    )
    for row_number, values in enumerate(summary_rows, 4):
        _append_row(summary, row_number, values)

    for sheet_name in (
        "通用工程类-实体工程量清单-东塔",
        "通用工程类-实体工程量清单-西塔",
    ):
        sheet = workbook[sheet_name]
        _append_row(sheet, 3, _ENTITY_HEADERS)
        _append_row(
            sheet,
            4,
            ["section", "root", "Y", "1", "section-unique", None, "1", None, "1", "灯具", "分类", "页签", None, None, None, "200", None],
        )
        _append_row(
            sheet,
            5,
            ["item", "section", "Y", "1-1", "item-unique", None, "1", None, "1-1", "LED灯", "10W", "清单项", "套", "2", "100", "200", ""],
        )

    opening = workbook["开办费清单"]
    _append_row(opening, 3, _GENERIC_HEADERS[:11] + ("不含税合价（元）", "说明"))
    _append_row(opening, 4, ["1", "section", "root", "Y", "1", "开办费", "", "页签", None, None, None, "10", None])
    _append_row(opening, 5, ["2", "item", "section", "Y", "2", "深化设计", "设计费", "清单项", "项", "1", "10", "10", ""])

    appendix_headers = ["序号", "名称", "计算规则", "单位", "数量", "不含税单价（元）", "不含税合价（元）", "备注"]
    for sheet_name in _SHEETS[7:13]:
        sheet = workbook[sheet_name]
        sheet["A1"] = sheet_name
        _append_row(sheet, 3, appendix_headers)
    for sheet_name, item_name in (
        ("附表01-安全文明施工费", "现场消防"),
        ("附表06-其他", "高空作业措施"),
    ):
        sheet = workbook[sheet_name]
        _append_row(sheet, 4, ["一", "分类", None, None, None, None, "2"])
        _append_row(sheet, 5, ["1", item_name, "", "项", "1", "2", "2", ""])

    management = workbook["照管费清单"]
    management["A1"] = "照管费清单"
    _append_row(management, 3, _GENERIC_HEADERS[:10] + ("照管费率（%）", "不含税合价（元）", "说明"))

    provisional = workbook["暂列清单"]
    provisional["A1"] = "暂列清单"
    _append_row(provisional, 3, _GENERIC_HEADERS[:8] + ("不含税延付金额（元）", "费率（%）", "不含税暂列金额（元）", "说明"))

    daywork = workbook["计日工工程量清单"]
    _append_row(daywork, 3, _GENERIC_HEADERS)
    _append_row(daywork, 4, [None] * 10 + ["指导价（元）", "X%", "价修后综合单价（元）"])
    daywork["L5"] = "指导价修正系数"
    _append_row(daywork, 6, ["1", "section", "root", "Y", "一", "人工", "", "页签", None, None, None, None, None, "10", None])
    _append_row(daywork, 7, ["2", "item", "section", "Y", "1", "杂工", "", "清单项", "工日", "1", "10", "0", "10", "10", ""])

    analysis_headers = ["项目编码", "项目名称", "项目特征描述", "计量单位", "不含税金额（元）"] + [None] * 9 + ["备注"]
    analysis_subheaders = [None] * 4 + ["综合单价", "人工费", "材料费", "其中主材", "主材损耗率", "损耗金额", "辅材", "机械费", "管理费", "利润"]
    for sheet_name in _SHEETS[16:]:
        sheet = workbook[sheet_name]
        _append_row(sheet, 3, analysis_headers)
        _append_row(sheet, 4, analysis_subheaders)
        _append_row(sheet, 5, ["1-1", "LED灯", "10W", "套", "100", "20", "50", "45", "0.1", "5", "0", "10", "10", "10", ""])

    workbook.save(path)
    return path.read_bytes()


def _boq_spec(settings, tmp_path):
    settings.MEDIA_ROOT = tmp_path / "media"
    user = get_user_model().objects.create_user(username="boq-reviewer")
    project = KnowledgeProject.objects.create(name="BOQ项目", slug="boq-project")
    document = Document.objects.create(
        project=project, name="工程量清单", kind=Document.Kind.PRICE_LIST
    )
    version = DocumentVersion.objects.create(document=document, version_label="v1")
    source = SourceFile.objects.create(
        document_version=version,
        role=SourceFile.Role.STRUCTURED_SOURCE,
        file=SimpleUploadedFile("boq.xlsx", _boq_xlsx_bytes(tmp_path)),
    )
    run_parser_for_source(source.pk, raise_errors=True)
    spec = BoqImportSpec(
        document_version=version,
        source_file=source,
        name="已验证清单模板",
        profile=BoqImportSpec.Profile.CRLAND_GENERAL_V1,
        project_name="测试泛光照明工程",
        party_a_name="测试甲方有限公司",
        party_a_group="测试集团",
        created_by=user,
    )
    spec.full_clean()
    spec.save()
    return spec, user


@pytest.mark.django_db
def test_boq_import_persists_context_items_summaries_and_analysis(settings, tmp_path):
    spec, reviewer = _boq_spec(settings, tmp_path)

    run = import_boq_spec(spec.pk)

    assert run.status == run.Status.SUCCEEDED
    assert run.imported_sheet_count == 18
    assert run.imported_item_count == 6
    assert run.imported_summary_count == 11
    assert run.sheet_snapshots.filter(is_empty_template=True).count() == 6
    assert set(run.item_records.values_list("project_name", flat=True)) == {
        "测试泛光照明工程"
    }
    assert set(run.item_records.values_list("party_a_name", flat=True)) == {
        "测试甲方有限公司"
    }
    assert set(run.item_records.values_list("party_a_group", flat=True)) == {
        "测试集团"
    }
    entity = run.item_records.get(kind=BoqItemRecord.Kind.ENTITY, tower_name="东塔")
    assert entity.quantity == Decimal("2.00000000")
    assert entity.unit_price == Decimal("100.00000000")
    assert entity.labor_cost == Decimal("20.00000000")
    assert entity.analysis_evidence.source_anchor["row"] == 5
    tax_rate = run.summary_records.get(kind=BoqSummaryRecord.Kind.TAX_RATE)
    assert tax_rate.rate == Decimal("0.09000000")
    assert tax_rate.project_name == "测试泛光照明工程"
    assert tax_rate.party_a_name == "测试甲方有限公司"
    assert tax_rate.party_a_group == "测试集团"
    assert trust_boq_item_records(run.item_records.all(), reviewer=reviewer) == 6
    assert trust_boq_summary_records(run.summary_records.all(), reviewer=reviewer) == 11


@pytest.mark.django_db
def test_boq_import_fails_as_one_batch_when_project_context_disagrees(settings, tmp_path):
    spec, _ = _boq_spec(settings, tmp_path)
    wrong = BoqImportSpec.objects.create(
        document_version=spec.document_version,
        source_file=spec.source_file,
        name="错误项目名",
        profile=BoqImportSpec.Profile.CRLAND_GENERAL_V1,
        project_name="另一个项目",
        party_a_name="测试甲方有限公司",
        party_a_group="测试集团",
    )

    run = import_boq_spec(wrong.pk)

    assert run.status == run.Status.FAILED
    assert "1-封皮!D6" in run.error
    assert run.item_records.count() == 0
    assert run.summary_records.count() == 0


def test_boq_import_spec_requires_manually_confirmed_party_a_name():
    spec = BoqImportSpec(project_name="项目", party_a_name="", party_a_group="集团")

    with pytest.raises(ValidationError, match="甲方名称不能为空"):
        spec.clean()


def test_boq_import_spec_requires_party_a_group():
    spec = BoqImportSpec(
        project_name="项目",
        party_a_name="甲方有限公司",
        party_a_group="",
    )

    with pytest.raises(ValidationError, match="甲方归属不能为空"):
        spec.clean()
