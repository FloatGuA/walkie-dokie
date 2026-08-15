"""Strict BOQ import for the verified CR Land commercial-lighting BIFF8 sample."""

from __future__ import annotations

from collections import Counter
from decimal import ROUND_HALF_UP, Decimal
import io
from pathlib import Path
from typing import Any

from openpyxl.utils import column_index_from_string, get_column_letter
import xlrd

from .boq import (
    _decimal,
    _evidence_index,
    _money,
    _require_money_equal,
    _row_evidence,
    _text,
)
from .models import (
    BoqImportRun,
    BoqImportSpec,
    BoqItemRecord,
    BoqSheetSnapshot,
    BoqSummaryRecord,
    ParserRun,
)


_SHEET_PROFILES = {
    "封面": (BoqSheetSnapshot.Kind.COVER, None),
    "编制说明": (BoqSheetSnapshot.Kind.NOTES, None),
    "汇总表": (BoqSheetSnapshot.Kind.SUMMARY, 2),
    "开办费清单": (BoqSheetSnapshot.Kind.OPENING, 5),
    "实体工程清单": (BoqSheetSnapshot.Kind.ENTITY, 2),
    "安全文明施工费": (BoqSheetSnapshot.Kind.OPENING_APPENDIX, 4),
    "综合单价分析表": (BoqSheetSnapshot.Kind.UNIT_PRICE_ANALYSIS, 2),
}

_SAFETY_OPENING_ROWS = frozenset({10, 11, 12, 13})


def _cell(sheet, row_number: int, column: str) -> Any:
    return sheet.cell_value(
        row_number - 1, column_index_from_string(column) - 1
    )


def _merged_value(sheet, row_number: int, column: str) -> Any:
    row_index = row_number - 1
    column_index = column_index_from_string(column) - 1
    value = sheet.cell_value(row_index, column_index)
    if value not in {None, ""}:
        return value
    for row_low, row_high, column_low, column_high in sheet.merged_cells:
        if row_low <= row_index < row_high and column_low <= column_index < column_high:
            return sheet.cell_value(row_low, column_low)
    return value


def _source_cells(sheet, row_number: int, columns: str) -> dict[str, dict[str, Any]]:
    cells: dict[str, dict[str, Any]] = {}
    row_index = row_number - 1
    for column in columns:
        column_index = column_index_from_string(column) - 1
        cell = sheet.cell(row_index, column_index)
        if cell.ctype in {xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK}:
            continue
        coordinate = f"{column}{row_number}"
        cells[coordinate] = {
            "raw_value": cell.value,
            "formula": None,
            "cached_value": None,
            "cell_type": cell.ctype,
            "formula_text_available": False,
        }
    return cells


def _require_headers(sheet, row_number: int, expected: dict[str, str]) -> None:
    for column, expected_value in expected.items():
        actual = _text(_cell(sheet, row_number, column))
        if actual != expected_value:
            raise ValueError(
                f"工作表 {sheet.name!r} {column}{row_number} "
                f"期望 {expected_value!r}，实际为 {actual!r}"
            )


def _required_decimal(sheet, row_number: int, column: str) -> Decimal:
    value = _decimal(
        _cell(sheet, row_number, column),
        field=f"{sheet.name}!{column}{row_number}",
    )
    if value is None:
        raise ValueError(f"工作表 {sheet.name!r} {column}{row_number} 数值为空")
    return value


def _optional_decimal(sheet, row_number: int, column: str) -> Decimal | None:
    return _decimal(
        _cell(sheet, row_number, column),
        field=f"{sheet.name}!{column}{row_number}",
    )


def _validate_sheet_profile(workbook) -> None:
    expected_names = list(_SHEET_PROFILES)
    if workbook.sheet_names() != expected_names:
        raise ValueError(
            "XLS 工作表集合或顺序与 crland_lighting_xls_v1 不一致；"
            "请创建新 profile，不允许模糊兼容"
        )
    hidden = [
        name
        for name in expected_names
        if workbook.sheet_by_name(name).visibility
    ]
    if hidden:
        raise ValueError(f"profile 要求全部工作表可见，但以下 sheet 被隐藏：{hidden}")


def _analysis_rows(workbook, evidence_by_row) -> dict[str, dict[str, Any]]:
    sheet_name = "综合单价分析表"
    sheet = workbook.sheet_by_name(sheet_name)
    _require_headers(
        sheet,
        2,
        {
            "A": "序号",
            "B": "项目名称",
            "C": "特征描述",
            "D": "计量\n单位",
            "E": "工程量",
            "F": "不含增值税综合单价a（a=b+c+h+i+j）",
            "G": "综合单价分析（元）",
            "P": "不含增值税汇总合价\n（元）",
        },
    )
    _require_headers(
        sheet,
        3,
        {
            "G": "人工费b",
            "H": "材料费c\n(c=d+f+g)",
            "I": "其中主材d",
            "J": "主材损耗率e",
            "K": "损耗金额f(f=dxe)",
            "L": "辅材g",
            "M": "机械费h",
            "N": "管理费i=（b+d）*费率",
            "O": "利润j=（b+c+h+i）*费率",
        },
    )
    rows: dict[str, dict[str, Any]] = {}
    for row_number in range(4, sheet.nrows + 1):
        unit = _text(_cell(sheet, row_number, "D"))
        if not unit:
            continue
        item_code = _text(_cell(sheet, row_number, "A"))
        if not item_code:
            raise ValueError(f"工作表 {sheet_name!r} 第 {row_number} 行缺少项目编码")
        if item_code in rows:
            raise ValueError(f"工作表 {sheet_name!r} 项目编码重复：{item_code}")
        row = {
            "row": row_number,
            "name": _text(_cell(sheet, row_number, "B")),
            "description": _text(_cell(sheet, row_number, "C")),
            "unit": unit,
            "quantity": _required_decimal(sheet, row_number, "E"),
            "unit_price": _required_decimal(sheet, row_number, "F"),
            "labor_cost": _optional_decimal(sheet, row_number, "G"),
            "material_cost": _optional_decimal(sheet, row_number, "H"),
            "main_material_cost": _optional_decimal(sheet, row_number, "I"),
            "main_material_loss_rate": _optional_decimal(sheet, row_number, "J"),
            "loss_cost": _optional_decimal(sheet, row_number, "K"),
            "auxiliary_material_cost": _optional_decimal(sheet, row_number, "L"),
            "machinery_cost": _optional_decimal(sheet, row_number, "M"),
            "management_cost": _optional_decimal(sheet, row_number, "N"),
            "profit": _optional_decimal(sheet, row_number, "O"),
            "total_price": _required_decimal(sheet, row_number, "P"),
            "source_cells": _source_cells(sheet, row_number, "ABCDEFGHIJKLMNOP"),
            "evidence": _row_evidence(evidence_by_row, sheet_name, row_number),
        }
        material_total = (
            (row["main_material_cost"] or Decimal("0"))
            + (row["loss_cost"] or Decimal("0"))
            + (row["auxiliary_material_cost"] or Decimal("0"))
        )
        _require_money_equal(
            row["material_cost"] or Decimal("0"),
            material_total,
            label=f"{sheet_name} {item_code} 材料费分解",
        )
        component_total = (
            (row["labor_cost"] or Decimal("0"))
            + (row["material_cost"] or Decimal("0"))
            + (row["machinery_cost"] or Decimal("0"))
            + (row["management_cost"] or Decimal("0"))
            + (row["profit"] or Decimal("0"))
        )
        _require_money_equal(
            row["unit_price"],
            component_total,
            label=f"{sheet_name} {item_code} 综合单价分解",
        )
        rows[item_code] = row
    return rows


def _entity_records(
    run: BoqImportRun,
    spec: BoqImportSpec,
    workbook,
    evidence_by_row,
) -> list[BoqItemRecord]:
    sheet_name = "实体工程清单"
    sheet = workbook.sheet_by_name(sheet_name)
    _require_headers(
        sheet,
        2,
        {
            "A": "序号",
            "B": "项目名称",
            "C": "特征描述",
            "D": "计量\n单位",
            "E": "工程量",
            "F": "不含增值税综合单价",
            "G": "不含增值税汇总合价\n（元）",
            "H": "备注",
        },
    )
    analysis_rows = _analysis_rows(workbook, evidence_by_row)
    consumed_analysis_codes: set[str] = set()
    records: list[BoqItemRecord] = []
    section_code = ""
    section_name = ""
    for row_number in range(3, sheet.nrows + 1):
        unit = _text(_cell(sheet, row_number, "D"))
        item_name = _text(_cell(sheet, row_number, "B"))
        if not unit:
            if item_name and item_name != "合计":
                section_code = _text(_cell(sheet, row_number, "A"))
                section_name = item_name
            continue
        item_code = _text(_cell(sheet, row_number, "A"))
        description = _text(_cell(sheet, row_number, "C"))
        if not item_code or not item_name:
            raise ValueError(f"工作表 {sheet_name!r} 第 {row_number} 行缺少编码或名称")
        analysis = analysis_rows.get(item_code)
        if analysis is None:
            raise ValueError(f"清单项 {item_code!r} 没有对应综合单价分析")
        quantity = _required_decimal(sheet, row_number, "E")
        unit_price = _required_decimal(sheet, row_number, "F")
        total_price = _required_decimal(sheet, row_number, "G")
        if (
            item_name != analysis["name"]
            or description != analysis["description"]
            or unit != analysis["unit"]
            or quantity != analysis["quantity"]
            or unit_price != analysis["unit_price"]
            or total_price != analysis["total_price"]
        ):
            raise ValueError(
                f"清单项 {item_code!r} 与综合单价分析的名称、特征、单位、数量或金额不一致"
            )
        expected_total = (quantity * unit_price).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        )
        _require_money_equal(
            total_price,
            expected_total,
            label=f"{sheet_name} {item_code} 数量乘单价",
        )
        consumed_analysis_codes.add(item_code)
        records.append(
            BoqItemRecord(
                import_run=run,
                document_version=spec.document_version,
                project_name=spec.project_name.strip(),
                party_a_name=spec.party_a_name.strip(),
                party_a_group=spec.party_a_group.strip(),
                kind=BoqItemRecord.Kind.ENTITY,
                section_code=section_code,
                section_name=section_name,
                item_code=item_code,
                item_name=item_name,
                item_description=description,
                unit=unit,
                quantity=quantity,
                unit_price=unit_price,
                total_price=total_price,
                labor_cost=analysis["labor_cost"],
                material_cost=analysis["material_cost"],
                main_material_cost=analysis["main_material_cost"],
                main_material_loss_rate=analysis["main_material_loss_rate"],
                loss_cost=analysis["loss_cost"],
                auxiliary_material_cost=analysis["auxiliary_material_cost"],
                machinery_cost=analysis["machinery_cost"],
                management_cost=analysis["management_cost"],
                profit=analysis["profit"],
                notes=_text(_cell(sheet, row_number, "H")),
                source_sheet=sheet_name,
                source_row=row_number,
                source_cells=_source_cells(sheet, row_number, "ABCDEFGH"),
                row_evidence=_row_evidence(evidence_by_row, sheet_name, row_number),
                header_evidence=_row_evidence(evidence_by_row, sheet_name, 2),
                analysis_source_sheet="综合单价分析表",
                analysis_source_row=analysis["row"],
                analysis_source_cells=analysis["source_cells"],
                analysis_evidence=analysis["evidence"],
                extensions={"formula_text_available": False},
            )
        )
    if consumed_analysis_codes != set(analysis_rows):
        unused = sorted(set(analysis_rows) - consumed_analysis_codes)
        raise ValueError(f"综合单价分析存在无对应清单项的编码：{unused}")
    return records


def _opening_records(
    run: BoqImportRun,
    spec: BoqImportSpec,
    workbook,
    evidence_by_row,
) -> tuple[list[BoqItemRecord], Decimal]:
    sheet_name = "开办费清单"
    sheet = workbook.sheet_by_name(sheet_name)
    _require_headers(
        sheet,
        5,
        {
            "A": "项目编码",
            "B": "项目名称",
            "C": "工作内容及包含范围",
            "D": "计量单位",
            "E": "工程量",
            "F": "不含税单价（元）",
            "G": "不含税合价（元）",
            "H": "说明",
        },
    )
    records: list[BoqItemRecord] = []
    omitted_safety_total = Decimal("0")
    for row_number in range(6, sheet.nrows + 1):
        item_name = _text(_merged_value(sheet, row_number, "B"))
        if item_name == "合计":
            continue
        unit = _text(_cell(sheet, row_number, "D"))
        if not unit:
            continue
        if row_number in _SAFETY_OPENING_ROWS:
            omitted_safety_total += _required_decimal(sheet, row_number, "G")
            continue
        if not item_name:
            raise ValueError(f"工作表 {sheet_name!r} 第 {row_number} 行缺少项目名称")
        records.append(
            BoqItemRecord(
                import_run=run,
                document_version=spec.document_version,
                project_name=spec.project_name.strip(),
                party_a_name=spec.party_a_name.strip(),
                party_a_group=spec.party_a_group.strip(),
                kind=BoqItemRecord.Kind.OPENING,
                item_code=_text(_merged_value(sheet, row_number, "A")),
                item_name=item_name,
                item_description=_text(_merged_value(sheet, row_number, "C")),
                unit=unit,
                quantity=_required_decimal(sheet, row_number, "E"),
                unit_price=_required_decimal(sheet, row_number, "F"),
                total_price=_required_decimal(sheet, row_number, "G"),
                notes=_text(_merged_value(sheet, row_number, "H")),
                source_sheet=sheet_name,
                source_row=row_number,
                source_cells=_source_cells(sheet, row_number, "ABCDEFGH"),
                row_evidence=_row_evidence(evidence_by_row, sheet_name, row_number),
                header_evidence=_row_evidence(evidence_by_row, sheet_name, 5),
                extensions={"formula_text_available": False},
            )
        )
    opening_total = _required_decimal(sheet, 41, "G")
    return records, opening_total - omitted_safety_total


def _safety_records(
    run: BoqImportRun,
    spec: BoqImportSpec,
    workbook,
    evidence_by_row,
) -> tuple[list[BoqItemRecord], Decimal]:
    sheet_name = "安全文明施工费"
    sheet = workbook.sheet_by_name(sheet_name)
    _require_headers(
        sheet,
        4,
        {
            "A": "序号",
            "B": "名称",
            "C": "计算规则",
            "D": "单位",
            "E": "数量",
            "F": "不含税单价(元)",
            "G": "合价(元)",
            "H": "备注",
        },
    )
    records: list[BoqItemRecord] = []
    section_code = ""
    section_name = ""
    for row_number in range(5, sheet.nrows + 1):
        item_name = _text(_cell(sheet, row_number, "B"))
        unit = _text(_cell(sheet, row_number, "D"))
        if item_name in {"合计", "合计："}:
            continue
        if not unit:
            if item_name:
                section_code = _text(_cell(sheet, row_number, "A"))
                section_name = item_name
            continue
        if not item_name:
            raise ValueError(f"工作表 {sheet_name!r} 第 {row_number} 行缺少项目名称")
        records.append(
            BoqItemRecord(
                import_run=run,
                document_version=spec.document_version,
                project_name=spec.project_name.strip(),
                party_a_name=spec.party_a_name.strip(),
                party_a_group=spec.party_a_group.strip(),
                kind=BoqItemRecord.Kind.OPENING_APPENDIX,
                section_code=section_code,
                section_name=section_name,
                item_code=_text(_cell(sheet, row_number, "A")),
                item_name=item_name,
                item_description=_text(_cell(sheet, row_number, "C")),
                unit=unit,
                quantity=_required_decimal(sheet, row_number, "E"),
                unit_price=_required_decimal(sheet, row_number, "F"),
                total_price=_required_decimal(sheet, row_number, "G"),
                notes=_text(_cell(sheet, row_number, "H")),
                source_sheet=sheet_name,
                source_row=row_number,
                source_cells=_source_cells(sheet, row_number, "ABCDEFGH"),
                row_evidence=_row_evidence(evidence_by_row, sheet_name, row_number),
                header_evidence=_row_evidence(evidence_by_row, sheet_name, 4),
                extensions={"formula_text_available": False},
            )
        )
    stated_total = _required_decimal(sheet, 27, "G")
    calculated_total = sum(
        (_money(record.total_price) for record in records), Decimal("0")
    )
    _require_money_equal(
        stated_total,
        calculated_total,
        label=f"{sheet_name} 叶子明细合计",
    )
    return records, stated_total


def _summary_records(
    run: BoqImportRun,
    spec: BoqImportSpec,
    workbook,
    evidence_by_row,
) -> list[BoqSummaryRecord]:
    summary_name = "汇总表"
    summary = workbook.sheet_by_name(summary_name)
    _require_headers(
        summary,
        2,
        {"A": "序号", "B": "名称", "C": "单位", "D": "金额（元）", "E": "备注"},
    )
    cover_name = "封面"
    cover = workbook.sheet_by_name(cover_name)
    definitions = (
        (
            summary_name,
            4,
            BoqSummaryRecord.Kind.COMPONENT,
            "开办费清单",
            "D",
            False,
        ),
        (
            summary_name,
            5,
            BoqSummaryRecord.Kind.COMPONENT,
            "实体工程清单",
            "D",
            False,
        ),
        (
            cover_name,
            12,
            BoqSummaryRecord.Kind.PRE_TAX_TOTAL,
            "合同不含增值税合价",
            "E",
            False,
        ),
        (
            cover_name,
            16,
            BoqSummaryRecord.Kind.TAX_RATE,
            "增值税税率",
            "E",
            True,
        ),
        (
            cover_name,
            18,
            BoqSummaryRecord.Kind.TAX_AMOUNT,
            "增值税合价",
            "E",
            False,
        ),
        (
            cover_name,
            22,
            BoqSummaryRecord.Kind.TAX_INCLUDED_TOTAL,
            "合同含增值税合价",
            "E",
            False,
        ),
    )
    records: list[BoqSummaryRecord] = []
    for sheet_name, row_number, kind, label, column, is_rate in definitions:
        sheet = workbook.sheet_by_name(sheet_name)
        value = _required_decimal(sheet, row_number, column)
        records.append(
            BoqSummaryRecord(
                import_run=run,
                document_version=spec.document_version,
                project_name=spec.project_name.strip(),
                party_a_name=spec.party_a_name.strip(),
                party_a_group=spec.party_a_group.strip(),
                kind=kind,
                summary_code=(
                    _text(_cell(sheet, row_number, "A"))
                    if sheet_name == summary_name
                    else ""
                ),
                summary_name=label,
                amount=None if is_rate else value,
                rate=value if is_rate else None,
                tax_included=kind == BoqSummaryRecord.Kind.TAX_INCLUDED_TOTAL,
                source_sheet=sheet_name,
                source_row=row_number,
                source_cells=_source_cells(sheet, row_number, "ABCDE"),
                row_evidence=_row_evidence(evidence_by_row, sheet_name, row_number),
                header_evidence=_row_evidence(
                    evidence_by_row,
                    sheet_name,
                    2 if sheet_name == summary_name else 2,
                ),
            )
        )
    return records


def _validate_totals(
    workbook,
    item_records: list[BoqItemRecord],
    summary_records: list[BoqSummaryRecord],
    opening_non_safety_total: Decimal,
    safety_total: Decimal,
) -> None:
    by_kind = {record.kind: record for record in summary_records}
    components = {
        record.summary_name: record
        for record in summary_records
        if record.kind == BoqSummaryRecord.Kind.COMPONENT
    }
    opening_total = components["开办费清单"].amount
    entity_total = components["实体工程清单"].amount
    if opening_total is None or entity_total is None:
        raise ValueError("开办费或实体工程汇总为空")
    calculated_opening_non_safety = sum(
        (
            _money(record.total_price)
            for record in item_records
            if record.kind == BoqItemRecord.Kind.OPENING
        ),
        Decimal("0"),
    )
    _require_money_equal(
        opening_non_safety_total,
        calculated_opening_non_safety,
        label="开办费主表非安全文明施工费明细",
    )
    _require_money_equal(
        opening_total,
        opening_non_safety_total + safety_total,
        label="开办费清单汇总",
    )
    calculated_entity = sum(
        (
            _money(record.total_price)
            for record in item_records
            if record.kind == BoqItemRecord.Kind.ENTITY
        ),
        Decimal("0"),
    )
    _require_money_equal(
        entity_total,
        calculated_entity,
        label="实体工程清单汇总",
    )
    pre_tax = by_kind[BoqSummaryRecord.Kind.PRE_TAX_TOTAL].amount
    tax_rate = by_kind[BoqSummaryRecord.Kind.TAX_RATE].rate
    tax_amount = by_kind[BoqSummaryRecord.Kind.TAX_AMOUNT].amount
    tax_included = by_kind[BoqSummaryRecord.Kind.TAX_INCLUDED_TOTAL].amount
    if None in {pre_tax, tax_rate, tax_amount, tax_included}:
        raise ValueError("封面不含税总价、税率、税额或含税总价为空")
    _require_money_equal(
        pre_tax,
        opening_total + entity_total,
        label="不含增值税合价",
    )
    calculated_tax = (pre_tax * tax_rate).quantize(
        Decimal("0.01"), rounding=ROUND_HALF_UP
    )
    _require_money_equal(tax_amount, calculated_tax, label="增值税合价")
    _require_money_equal(
        tax_included,
        _money(pre_tax) + _money(tax_amount),
        label="含增值税合价",
    )

    summary = workbook.sheet_by_name("汇总表")
    _require_money_equal(
        _required_decimal(summary, 6, "D"),
        tax_amount,
        label="汇总表税金",
    )
    _require_money_equal(
        _required_decimal(summary, 7, "D"),
        tax_included,
        label="汇总表含税合计",
    )


def _sheet_snapshots(
    run: BoqImportRun,
    workbook,
    item_records: list[BoqItemRecord],
    summary_records: list[BoqSummaryRecord],
) -> list[BoqSheetSnapshot]:
    imported_counts = Counter(record.source_sheet for record in item_records)
    imported_counts.update(
        record.analysis_source_sheet
        for record in item_records
        if record.analysis_source_sheet
    )
    imported_counts.update(record.source_sheet for record in summary_records)
    snapshots: list[BoqSheetSnapshot] = []
    for sheet_name, (kind, header_row) in _SHEET_PROFILES.items():
        sheet = workbook.sheet_by_name(sheet_name)
        nonempty_rows = 0
        for row_index in range(sheet.nrows):
            nonempty_rows += int(
                any(
                    sheet.cell_type(row_index, column_index)
                    not in {xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK}
                    for column_index in range(sheet.row_len(row_index))
                )
            )
        snapshots.append(
            BoqSheetSnapshot(
                import_run=run,
                source_sheet=sheet_name,
                kind=kind,
                header_row=header_row,
                nonempty_row_count=nonempty_rows,
                formula_count=0,
                imported_record_count=imported_counts[sheet_name],
                is_empty_template=False,
                metadata={
                    "max_row": sheet.nrows,
                    "max_column": sheet.ncols,
                    "hidden_columns": [
                        get_column_letter(column_index + 1)
                        for column_index, info in sheet.colinfo_map.items()
                        if info.hidden
                    ],
                    "hidden_rows": [
                        row_index + 1
                        for row_index, info in sheet.rowinfo_map.items()
                        if info.hidden
                    ],
                    "formula_count_known": False,
                },
            )
        )
    return snapshots


def build_crland_lighting_xls_import(
    run: BoqImportRun,
    spec: BoqImportSpec,
    parser_run: ParserRun,
) -> tuple[
    list[BoqSheetSnapshot],
    list[BoqItemRecord],
    list[BoqSummaryRecord],
    list[str],
]:
    parser_log = io.StringIO()
    workbook = xlrd.open_workbook(
        Path(spec.source_file.file.path),
        on_demand=True,
        formatting_info=True,
        logfile=parser_log,
    )
    try:
        _validate_sheet_profile(workbook)
        cover = workbook.sheet_by_name("封面")
        project_name = _text(_cell(cover, 9, "E"))
        party_a_name = _text(_cell(cover, 7, "E"))
        if project_name != spec.project_name.strip():
            raise ValueError(
                f"导入配置项目名称 {spec.project_name!r} 与 封面!E9 "
                f"{project_name!r} 不一致"
            )
        if party_a_name != spec.party_a_name.strip():
            raise ValueError(
                f"导入配置甲方名称 {spec.party_a_name!r} 与 封面!E7 "
                f"{party_a_name!r} 不一致"
            )

        evidence_by_row = _evidence_index(parser_run)
        entity_records = _entity_records(
            run, spec, workbook, evidence_by_row
        )
        opening_records, opening_non_safety_total = _opening_records(
            run, spec, workbook, evidence_by_row
        )
        safety_records, safety_total = _safety_records(
            run, spec, workbook, evidence_by_row
        )
        item_records = entity_records + opening_records + safety_records
        summary_records = _summary_records(
            run, spec, workbook, evidence_by_row
        )
        _validate_totals(
            workbook,
            item_records,
            summary_records,
            opening_non_safety_total,
            safety_total,
        )
        snapshots = _sheet_snapshots(
            run, workbook, item_records, summary_records
        )
        zero_opening_count = sum(
            record.unit_price == Decimal("0")
            for record in opening_records
        )
        warnings = [
            "XLS 公式文本无法完整保留；已用叶子明细、成本分解和税额独立闭合，仍只进入 Staging",
            (
                f"开办费有 {zero_opening_count} 条零价记录；"
                "原文说明未填价费用视为已包含在其他单价或合价中"
            ),
        ]
        parser_messages = [
            line.strip()
            for line in parser_log.getvalue().splitlines()
            if line.strip()
        ]
        if parser_messages:
            warnings.append(
                f"XLS parser 报告 {len(parser_messages)} 个未识别公式函数记录；未将其视为受控重算"
            )
        return snapshots, item_records, summary_records, warnings
    finally:
        workbook.release_resources()
