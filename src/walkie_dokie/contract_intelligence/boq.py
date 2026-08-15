"""Deterministic ingestion for explicitly supported bill-of-quantities workbooks."""

from __future__ import annotations

from collections import Counter
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .models import (
    BoqImportRun,
    BoqImportSpec,
    BoqItemRecord,
    BoqSheetSnapshot,
    BoqSummaryRecord,
    EvidenceUnit,
    ParserRun,
)


_SHEET_PROFILES = {
    "招标清单列表": (BoqSheetSnapshot.Kind.MANIFEST, 1),
    "1-封皮": (BoqSheetSnapshot.Kind.COVER, None),
    "2-编制说明": (BoqSheetSnapshot.Kind.NOTES, None),
    "3-汇总表": (BoqSheetSnapshot.Kind.SUMMARY, 3),
    "通用工程类-实体工程量清单-东塔": (BoqSheetSnapshot.Kind.ENTITY, 3),
    "通用工程类-实体工程量清单-西塔": (BoqSheetSnapshot.Kind.ENTITY, 3),
    "开办费清单": (BoqSheetSnapshot.Kind.OPENING, 3),
    "附表01-安全文明施工费": (BoqSheetSnapshot.Kind.OPENING_APPENDIX, 3),
    "附表02-垂直运输机械": (BoqSheetSnapshot.Kind.OPENING_APPENDIX, 3),
    "附表03-脚手架": (BoqSheetSnapshot.Kind.OPENING_APPENDIX, 3),
    "附表04-临时使用永久电梯": (BoqSheetSnapshot.Kind.OPENING_APPENDIX, 3),
    "附表05-试验与检验": (BoqSheetSnapshot.Kind.OPENING_APPENDIX, 3),
    "附表06-其他": (BoqSheetSnapshot.Kind.OPENING_APPENDIX, 3),
    "照管费清单": (BoqSheetSnapshot.Kind.MANAGEMENT, 3),
    "计日工工程量清单": (BoqSheetSnapshot.Kind.DAYWORK, 3),
    "暂列清单": (BoqSheetSnapshot.Kind.PROVISIONAL, 3),
    "通用工程类-实体工程量清单-东塔-综合单价分析表": (
        BoqSheetSnapshot.Kind.UNIT_PRICE_ANALYSIS,
        3,
    ),
    "通用工程类-实体工程量清单-西塔-综合单价分析表": (
        BoqSheetSnapshot.Kind.UNIT_PRICE_ANALYSIS,
        3,
    ),
}

_ENTITY_SHEETS = {
    "通用工程类-实体工程量清单-东塔": "东塔",
    "通用工程类-实体工程量清单-西塔": "西塔",
}

_ANALYSIS_SHEETS = {
    "东塔": "通用工程类-实体工程量清单-东塔-综合单价分析表",
    "西塔": "通用工程类-实体工程量清单-西塔-综合单价分析表",
}

_OPENING_APPENDIX_SHEETS = (
    "附表01-安全文明施工费",
    "附表02-垂直运输机械",
    "附表03-脚手架",
    "附表04-临时使用永久电梯",
    "附表05-试验与检验",
    "附表06-其他",
)


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _decimal(value: Any, *, field: str) -> Decimal | None:
    if value in {None, ""}:
        return None
    cleaned = str(value).strip().replace(",", "").replace("￥", "").replace("¥", "")
    try:
        return Decimal(cleaned)
    except InvalidOperation as exc:
        raise ValueError(f"{field} 不是合法十进制数：{value!r}") from exc


def _money(value: Decimal | None) -> Decimal:
    if value is None:
        return Decimal("0.00")
    return value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _require_money_equal(actual: Decimal, expected: Decimal, *, label: str) -> None:
    if _money(actual) != _money(expected):
        raise ValueError(
            f"{label} 不一致：实际 {_money(actual)}，按明细重算 {_money(expected)}"
        )


def _cell_value(formula_sheet, value_sheet, coordinate: str) -> Any:
    raw_value = formula_sheet[coordinate].value
    if isinstance(raw_value, str) and raw_value.startswith("="):
        cached_value = value_sheet[coordinate].value
        if cached_value is None:
            raise ValueError(
                f"工作表 {formula_sheet.title!r} 单元格 {coordinate} 公式没有缓存值"
            )
        return cached_value
    return raw_value


def _source_cells(formula_sheet, value_sheet, row_number: int, columns: str) -> dict:
    cells = {}
    for column in columns:
        coordinate = f"{column}{row_number}"
        formula_cell = formula_sheet[coordinate]
        value = _cell_value(formula_sheet, value_sheet, coordinate)
        if formula_cell.value is None and value is None:
            continue
        is_formula = isinstance(formula_cell.value, str) and formula_cell.value.startswith("=")
        cells[coordinate] = {
            "raw_value": None if formula_cell.value is None else str(formula_cell.value),
            "formula": formula_cell.value if is_formula else None,
            "cached_value": None if not is_formula or value is None else str(value),
            "number_format": formula_cell.number_format,
        }
    return cells


def _require_headers(value_sheet, row_number: int, expected: dict[str, str]) -> None:
    for column, expected_value in expected.items():
        actual = _text(value_sheet[f"{column}{row_number}"].value)
        if actual != expected_value:
            raise ValueError(
                f"工作表 {value_sheet.title!r} {column}{row_number} "
                f"期望 {expected_value!r}，实际为 {actual!r}"
            )


def _evidence_index(parser_run: ParserRun) -> dict[tuple[str, int], EvidenceUnit]:
    result = {}
    for evidence in parser_run.evidence_units.filter(kind="spreadsheet_row"):
        key = (
            _text(evidence.source_anchor.get("sheet")),
            int(evidence.source_anchor.get("row")),
        )
        if key in result:
            raise ValueError(f"存在重复 Evidence Unit：{key[0]!r} 第 {key[1]} 行")
        result[key] = evidence
    return result


def _row_evidence(
    evidence_by_row: dict[tuple[str, int], EvidenceUnit],
    sheet_name: str,
    row_number: int,
) -> EvidenceUnit:
    evidence = evidence_by_row.get((sheet_name, row_number))
    if evidence is None:
        raise ValueError(f"工作表 {sheet_name!r} 第 {row_number} 行没有 Evidence Unit")
    return evidence


def _validate_workbook_profile(formula_workbook, value_workbook) -> None:
    expected_names = list(_SHEET_PROFILES)
    if formula_workbook.sheetnames != expected_names:
        raise ValueError(
            "XLSX 工作表集合或顺序与 crland_general_v1 不一致；"
            "请创建新 profile，不允许模糊兼容"
        )
    if value_workbook.sheetnames != expected_names:
        raise ValueError("公式工作簿与缓存值工作簿的 sheet 不一致")
    for sheet_name in expected_names:
        formula_sheet = formula_workbook[sheet_name]
        value_sheet = value_workbook[sheet_name]
        for row in formula_sheet.iter_rows():
            for cell in row:
                if not (isinstance(cell.value, str) and cell.value.startswith("=")):
                    continue
                if value_sheet[cell.coordinate].value is None:
                    raise ValueError(
                        f"工作表 {sheet_name!r} 单元格 {cell.coordinate} "
                        "公式没有缓存值"
                    )


def _manifest_warnings(value_workbook) -> list[str]:
    manifest_sheet = value_workbook["招标清单列表"]
    listed_names = {
        _text(manifest_sheet[f"B{row_number}"].value)
        for row_number in range(2, manifest_sheet.max_row + 1)
        if _text(manifest_sheet[f"B{row_number}"].value)
    }
    actual_names = set(value_workbook.sheetnames)
    missing = sorted(listed_names - actual_names)
    unlisted = sorted(actual_names - listed_names)
    warnings = []
    if missing:
        warnings.append(f"清单目录列出但实际不存在的 sheet：{missing}")
    if unlisted:
        warnings.append(f"实际存在但清单目录未列出的 sheet：{unlisted}")
    return warnings


def _analysis_rows(
    formula_workbook,
    value_workbook,
    evidence_by_row,
    tower_name: str,
) -> dict[str, dict]:
    sheet_name = _ANALYSIS_SHEETS[tower_name]
    formula_sheet = formula_workbook[sheet_name]
    value_sheet = value_workbook[sheet_name]
    _require_headers(
        value_sheet,
        3,
        {
            "A": "项目编码",
            "B": "项目名称",
            "C": "项目特征描述",
            "D": "计量单位",
            "E": "不含税金额（元）",
            "O": "备注",
        },
    )
    _require_headers(
        value_sheet,
        4,
        {
            "E": "综合单价",
            "F": "人工费",
            "G": "材料费",
            "H": "其中主材",
            "I": "主材损耗率",
            "J": "损耗金额",
            "K": "辅材",
            "L": "机械费",
            "M": "管理费",
            "N": "利润",
        },
    )
    result = {}
    for row_number in range(5, value_sheet.max_row + 1):
        if not _text(value_sheet[f"D{row_number}"].value):
            continue
        item_code = _text(value_sheet[f"A{row_number}"].value)
        if not item_code:
            raise ValueError(f"工作表 {sheet_name!r} 第 {row_number} 行缺少项目编码")
        if item_code in result:
            raise ValueError(f"工作表 {sheet_name!r} 项目编码重复：{item_code}")
        result[item_code] = {
            "row": row_number,
            "name": _text(value_sheet[f"B{row_number}"].value),
            "description": _text(value_sheet[f"C{row_number}"].value),
            "unit": _text(value_sheet[f"D{row_number}"].value),
            "unit_price": _decimal(
                _cell_value(formula_sheet, value_sheet, f"E{row_number}"),
                field=f"{sheet_name}!E{row_number}",
            ),
            "labor_cost": _decimal(
                _cell_value(formula_sheet, value_sheet, f"F{row_number}"),
                field=f"{sheet_name}!F{row_number}",
            ),
            "material_cost": _decimal(
                _cell_value(formula_sheet, value_sheet, f"G{row_number}"),
                field=f"{sheet_name}!G{row_number}",
            ),
            "main_material_cost": _decimal(
                _cell_value(formula_sheet, value_sheet, f"H{row_number}"),
                field=f"{sheet_name}!H{row_number}",
            ),
            "main_material_loss_rate": _decimal(
                _cell_value(formula_sheet, value_sheet, f"I{row_number}"),
                field=f"{sheet_name}!I{row_number}",
            ),
            "loss_cost": _decimal(
                _cell_value(formula_sheet, value_sheet, f"J{row_number}"),
                field=f"{sheet_name}!J{row_number}",
            ),
            "auxiliary_material_cost": _decimal(
                _cell_value(formula_sheet, value_sheet, f"K{row_number}"),
                field=f"{sheet_name}!K{row_number}",
            ),
            "machinery_cost": _decimal(
                _cell_value(formula_sheet, value_sheet, f"L{row_number}"),
                field=f"{sheet_name}!L{row_number}",
            ),
            "management_cost": _decimal(
                _cell_value(formula_sheet, value_sheet, f"M{row_number}"),
                field=f"{sheet_name}!M{row_number}",
            ),
            "profit": _decimal(
                _cell_value(formula_sheet, value_sheet, f"N{row_number}"),
                field=f"{sheet_name}!N{row_number}",
            ),
            "notes": _text(value_sheet[f"O{row_number}"].value),
            "source_cells": _source_cells(
                formula_sheet, value_sheet, row_number, "ABCDEFGHIJKLMNO"
            ),
            "evidence": _row_evidence(evidence_by_row, sheet_name, row_number),
        }
    return result


def _entity_records(
    run,
    spec,
    formula_workbook,
    value_workbook,
    evidence_by_row,
) -> list[BoqItemRecord]:
    records = []
    for sheet_name, tower_name in _ENTITY_SHEETS.items():
        formula_sheet = formula_workbook[sheet_name]
        value_sheet = value_workbook[sheet_name]
        _require_headers(
            value_sheet,
            3,
            {
                "A": "itemId",
                "B": "parentId",
                "C": "leafFlag",
                "D": "itemSortCode",
                "E": "唯一标识",
                "I": "项目编号",
                "J": "项目名称",
                "K": "项目特征描述",
                "L": "类别",
                "M": "计量单位",
                "N": "工程数量",
                "O": "不含增值税综合单价（元）",
                "P": "不含增值税汇总合价（元）",
                "Q": "备注",
            },
        )
        analysis_rows = _analysis_rows(
            formula_workbook, value_workbook, evidence_by_row, tower_name
        )
        consumed_analysis_codes = set()
        section_code = ""
        section_name = ""
        for row_number in range(4, value_sheet.max_row + 1):
            row_kind = _text(value_sheet[f"L{row_number}"].value)
            if row_kind == "页签":
                section_code = _text(value_sheet[f"I{row_number}"].value)
                section_name = _text(value_sheet[f"J{row_number}"].value)
                continue
            if row_kind != "清单项":
                raise ValueError(
                    f"工作表 {sheet_name!r} 第 {row_number} 行类别不受支持：{row_kind!r}"
                )
            if not section_name:
                raise ValueError(f"工作表 {sheet_name!r} 第 {row_number} 行没有所属页签")
            item_code = _text(value_sheet[f"I{row_number}"].value)
            item_name = _text(value_sheet[f"J{row_number}"].value)
            unit = _text(value_sheet[f"M{row_number}"].value)
            if not item_code or not item_name or not unit:
                raise ValueError(f"工作表 {sheet_name!r} 第 {row_number} 行缺少编码、名称或单位")
            analysis = analysis_rows.get(item_code)
            if analysis is None:
                raise ValueError(
                    f"{tower_name} 清单项 {item_code!r} 没有对应综合单价分析"
                )
            unit_price = _decimal(
                _cell_value(formula_sheet, value_sheet, f"O{row_number}"),
                field=f"{sheet_name}!O{row_number}",
            )
            if (
                item_name != analysis["name"]
                or unit != analysis["unit"]
                or unit_price != analysis["unit_price"]
            ):
                raise ValueError(
                    f"{tower_name} 清单项 {item_code!r} 与综合单价分析的名称、单位或单价不一致"
                )
            component_total = sum(
                (
                    analysis["labor_cost"] or Decimal("0"),
                    analysis["material_cost"] or Decimal("0"),
                    analysis["machinery_cost"] or Decimal("0"),
                    analysis["management_cost"] or Decimal("0"),
                    analysis["profit"] or Decimal("0"),
                ),
                Decimal("0"),
            )
            if unit_price is None or abs(unit_price - component_total) > Decimal(
                "0.00000001"
            ):
                raise ValueError(
                    f"{tower_name} 清单项 {item_code!r} 综合单价与成本分解不一致"
                )
            consumed_analysis_codes.add(item_code)
            records.append(
                BoqItemRecord(
                    import_run=run,
                    document_version=spec.document_version,
                    project_name=spec.project_name.strip(),
                    party_a_name=spec.party_a_name.strip(),
                    party_a_group=spec.party_a_group.strip(),
                    kind=BoqItemRecord.Kind.ENTITY,
                    tower_name=tower_name,
                    section_code=section_code,
                    section_name=section_name,
                    source_item_id=_text(value_sheet[f"A{row_number}"].value),
                    source_parent_id=_text(value_sheet[f"B{row_number}"].value),
                    source_unique_id=_text(value_sheet[f"E{row_number}"].value),
                    source_sort_code=_text(value_sheet[f"D{row_number}"].value),
                    item_code=item_code,
                    item_name=item_name,
                    item_description=_text(value_sheet[f"K{row_number}"].value),
                    unit=unit,
                    quantity=_decimal(
                        _cell_value(formula_sheet, value_sheet, f"N{row_number}"),
                        field=f"{sheet_name}!N{row_number}",
                    ),
                    unit_price=unit_price,
                    total_price=_decimal(
                        _cell_value(formula_sheet, value_sheet, f"P{row_number}"),
                        field=f"{sheet_name}!P{row_number}",
                    ),
                    labor_cost=analysis["labor_cost"],
                    material_cost=analysis["material_cost"],
                    main_material_cost=analysis["main_material_cost"],
                    main_material_loss_rate=analysis["main_material_loss_rate"],
                    loss_cost=analysis["loss_cost"],
                    auxiliary_material_cost=analysis["auxiliary_material_cost"],
                    machinery_cost=analysis["machinery_cost"],
                    management_cost=analysis["management_cost"],
                    profit=analysis["profit"],
                    notes=_text(value_sheet[f"Q{row_number}"].value),
                    extensions={
                        "source_leaf_flag": _text(value_sheet[f"C{row_number}"].value),
                        "source_list_id": _text(value_sheet[f"F{row_number}"].value),
                        "source_data_origin": _text(value_sheet[f"G{row_number}"].value),
                        "source_custom_import": _text(value_sheet[f"H{row_number}"].value),
                        "analysis_description": analysis["description"],
                        "analysis_notes": analysis["notes"],
                    },
                    source_sheet=sheet_name,
                    source_row=row_number,
                    source_cells=_source_cells(
                        formula_sheet, value_sheet, row_number, "ABCDEFGHIJKLMNOPQ"
                    ),
                    row_evidence=_row_evidence(evidence_by_row, sheet_name, row_number),
                    header_evidence=_row_evidence(evidence_by_row, sheet_name, 3),
                    analysis_source_sheet=_ANALYSIS_SHEETS[tower_name],
                    analysis_source_row=analysis["row"],
                    analysis_source_cells=analysis["source_cells"],
                    analysis_evidence=analysis["evidence"],
                )
            )
        if consumed_analysis_codes != set(analysis_rows):
            unused = sorted(set(analysis_rows) - consumed_analysis_codes)
            raise ValueError(f"{tower_name} 综合单价分析存在无对应清单项的编码：{unused}")
    return records


def _opening_records(
    run,
    spec,
    formula_workbook,
    value_workbook,
    evidence_by_row,
) -> list[BoqItemRecord]:
    sheet_name = "开办费清单"
    formula_sheet = formula_workbook[sheet_name]
    value_sheet = value_workbook[sheet_name]
    _require_headers(
        value_sheet,
        3,
        {
            "A": "id",
            "B": "itemId",
            "C": "parentId",
            "D": "leafFlag",
            "E": "项目编码",
            "F": "项目名称",
            "G": "项目特征描述",
            "H": "类别",
            "I": "计量单位",
            "J": "工程量",
            "K": "不含税单价（元）",
            "L": "不含税合价（元）",
            "M": "说明",
        },
    )
    records = []
    section_code = ""
    section_name = ""
    for row_number in range(4, value_sheet.max_row + 1):
        row_kind = _text(value_sheet[f"H{row_number}"].value)
        if row_kind == "页签":
            section_code = _text(value_sheet[f"E{row_number}"].value)
            section_name = _text(value_sheet[f"F{row_number}"].value)
            continue
        if row_kind != "清单项":
            raise ValueError(
                f"工作表 {sheet_name!r} 第 {row_number} 行类别不受支持：{row_kind!r}"
            )
        records.append(
            BoqItemRecord(
                import_run=run,
                document_version=spec.document_version,
                project_name=spec.project_name.strip(),
                party_a_name=spec.party_a_name.strip(),
                party_a_group=spec.party_a_group.strip(),
                kind=BoqItemRecord.Kind.OPENING,
                section_code=section_code,
                section_name=section_name,
                source_item_id=_text(value_sheet[f"B{row_number}"].value),
                source_parent_id=_text(value_sheet[f"C{row_number}"].value),
                source_unique_id=_text(value_sheet[f"A{row_number}"].value),
                item_code=_text(value_sheet[f"E{row_number}"].value),
                item_name=_text(value_sheet[f"F{row_number}"].value),
                item_description=_text(value_sheet[f"G{row_number}"].value),
                unit=_text(value_sheet[f"I{row_number}"].value),
                quantity=_decimal(
                    _cell_value(formula_sheet, value_sheet, f"J{row_number}"),
                    field=f"{sheet_name}!J{row_number}",
                ),
                unit_price=_decimal(
                    _cell_value(formula_sheet, value_sheet, f"K{row_number}"),
                    field=f"{sheet_name}!K{row_number}",
                ),
                total_price=_decimal(
                    _cell_value(formula_sheet, value_sheet, f"L{row_number}"),
                    field=f"{sheet_name}!L{row_number}",
                ),
                notes=_text(value_sheet[f"M{row_number}"].value),
                extensions={
                    "source_leaf_flag": _text(value_sheet[f"D{row_number}"].value)
                },
                source_sheet=sheet_name,
                source_row=row_number,
                source_cells=_source_cells(
                    formula_sheet, value_sheet, row_number, "ABCDEFGHIJKLM"
                ),
                row_evidence=_row_evidence(evidence_by_row, sheet_name, row_number),
                header_evidence=_row_evidence(evidence_by_row, sheet_name, 3),
            )
        )
    return records


def _opening_appendix_records(
    run,
    spec,
    formula_workbook,
    value_workbook,
    evidence_by_row,
) -> list[BoqItemRecord]:
    records = []
    for sheet_name in _OPENING_APPENDIX_SHEETS:
        formula_sheet = formula_workbook[sheet_name]
        value_sheet = value_workbook[sheet_name]
        section_code = ""
        section_name = ""
        for row_number in range(4, value_sheet.max_row + 1):
            unit = _text(value_sheet[f"D{row_number}"].value)
            if not unit:
                if _text(value_sheet[f"B{row_number}"].value):
                    section_code = _text(value_sheet[f"A{row_number}"].value)
                    section_name = _text(value_sheet[f"B{row_number}"].value)
                continue
            item_name = _text(value_sheet[f"B{row_number}"].value)
            if not item_name:
                raise ValueError(f"工作表 {sheet_name!r} 第 {row_number} 行缺少项目名称")
            records.append(
                BoqItemRecord(
                    import_run=run,
                    document_version=spec.document_version,
                    project_name=spec.project_name.strip(),
                    party_a_name=spec.party_a_name.strip(),
                    party_a_group=spec.party_a_group.strip(),
                    kind=BoqItemRecord.Kind.OPENING_APPENDIX,
                    section_code=section_code,
                    section_name=section_name,
                    item_code=_text(value_sheet[f"A{row_number}"].value),
                    item_name=item_name,
                    item_description=_text(value_sheet[f"C{row_number}"].value),
                    unit=unit,
                    quantity=_decimal(
                        _cell_value(formula_sheet, value_sheet, f"E{row_number}"),
                        field=f"{sheet_name}!E{row_number}",
                    ),
                    unit_price=_decimal(
                        _cell_value(formula_sheet, value_sheet, f"F{row_number}"),
                        field=f"{sheet_name}!F{row_number}",
                    ),
                    total_price=_decimal(
                        _cell_value(formula_sheet, value_sheet, f"G{row_number}"),
                        field=f"{sheet_name}!G{row_number}",
                    ),
                    notes=_text(value_sheet[f"H{row_number}"].value),
                    source_sheet=sheet_name,
                    source_row=row_number,
                    source_cells=_source_cells(
                        formula_sheet, value_sheet, row_number, "ABCDEFGH"
                    ),
                    row_evidence=_row_evidence(evidence_by_row, sheet_name, row_number),
                    header_evidence=_row_evidence(evidence_by_row, sheet_name, 3),
                )
            )
    return records


def _daywork_records(
    run,
    spec,
    formula_workbook,
    value_workbook,
    evidence_by_row,
) -> list[BoqItemRecord]:
    sheet_name = "计日工工程量清单"
    formula_sheet = formula_workbook[sheet_name]
    value_sheet = value_workbook[sheet_name]
    _require_headers(
        value_sheet,
        3,
        {
            "A": "id",
            "B": "itemId",
            "C": "parentId",
            "D": "leafFlag",
            "E": "项目编码",
            "F": "项目名称",
            "G": "项目特征描述",
            "H": "类别",
            "I": "计量单位",
            "J": "工程量",
            "K": "不含税单价（元）",
            "N": "不含税合价（元）",
            "O": "说明",
        },
    )
    _require_headers(
        value_sheet,
        4,
        {"K": "指导价（元）", "L": "X%", "M": "价修后综合单价（元）"},
    )
    records = []
    section_code = ""
    section_name = ""
    for row_number in range(6, value_sheet.max_row + 1):
        row_kind = _text(value_sheet[f"H{row_number}"].value)
        if row_kind == "页签":
            section_code = _text(value_sheet[f"E{row_number}"].value)
            section_name = _text(value_sheet[f"F{row_number}"].value)
            continue
        if row_kind != "清单项":
            raise ValueError(
                f"工作表 {sheet_name!r} 第 {row_number} 行类别不受支持：{row_kind!r}"
            )
        item_name = _text(value_sheet[f"F{row_number}"].value)
        if not item_name:
            raise ValueError(f"工作表 {sheet_name!r} 第 {row_number} 行缺少项目名称")
        records.append(
            BoqItemRecord(
                import_run=run,
                document_version=spec.document_version,
                project_name=spec.project_name.strip(),
                party_a_name=spec.party_a_name.strip(),
                party_a_group=spec.party_a_group.strip(),
                kind=BoqItemRecord.Kind.DAYWORK,
                section_code=section_code,
                section_name=section_name,
                source_item_id=_text(value_sheet[f"B{row_number}"].value),
                source_parent_id=_text(value_sheet[f"C{row_number}"].value),
                source_unique_id=_text(value_sheet[f"A{row_number}"].value),
                item_code=_text(value_sheet[f"E{row_number}"].value),
                item_name=item_name,
                item_description=_text(value_sheet[f"G{row_number}"].value),
                unit=_text(value_sheet[f"I{row_number}"].value),
                quantity=_decimal(
                    _cell_value(formula_sheet, value_sheet, f"J{row_number}"),
                    field=f"{sheet_name}!J{row_number}",
                ),
                unit_price=_decimal(
                    _cell_value(formula_sheet, value_sheet, f"M{row_number}"),
                    field=f"{sheet_name}!M{row_number}",
                ),
                total_price=_decimal(
                    _cell_value(formula_sheet, value_sheet, f"N{row_number}"),
                    field=f"{sheet_name}!N{row_number}",
                ),
                guide_price=_decimal(
                    _cell_value(formula_sheet, value_sheet, f"K{row_number}"),
                    field=f"{sheet_name}!K{row_number}",
                ),
                adjustment_rate=_decimal(
                    _cell_value(formula_sheet, value_sheet, f"L{row_number}"),
                    field=f"{sheet_name}!L{row_number}",
                ),
                notes=_text(value_sheet[f"O{row_number}"].value),
                extensions={
                    "source_leaf_flag": _text(value_sheet[f"D{row_number}"].value)
                },
                source_sheet=sheet_name,
                source_row=row_number,
                source_cells=_source_cells(
                    formula_sheet, value_sheet, row_number, "ABCDEFGHIJKLMNO"
                ),
                row_evidence=_row_evidence(evidence_by_row, sheet_name, row_number),
                header_evidence=_row_evidence(evidence_by_row, sheet_name, 3),
            )
        )
    return records


def _summary_records(
    run,
    spec,
    formula_workbook,
    value_workbook,
    evidence_by_row,
) -> list[BoqSummaryRecord]:
    records = []
    sheet_name = "3-汇总表"
    formula_sheet = formula_workbook[sheet_name]
    value_sheet = value_workbook[sheet_name]
    _require_headers(
        value_sheet,
        3,
        {"A": "项目序号", "B": "项目", "C": "不含增值税合价", "D": "备注"},
    )
    for row_number in range(4, 12):
        summary_name = _text(value_sheet[f"B{row_number}"].value)
        if not summary_name:
            raise ValueError(f"工作表 {sheet_name!r} 第 {row_number} 行缺少汇总名称")
        if row_number == 11:
            kind = BoqSummaryRecord.Kind.PRE_TAX_TOTAL
        elif row_number in {6, 7}:
            kind = BoqSummaryRecord.Kind.SUBTOTAL
        else:
            kind = BoqSummaryRecord.Kind.COMPONENT
        records.append(
            BoqSummaryRecord(
                import_run=run,
                document_version=spec.document_version,
                project_name=spec.project_name.strip(),
                party_a_name=spec.party_a_name.strip(),
                party_a_group=spec.party_a_group.strip(),
                kind=kind,
                summary_code=_text(value_sheet[f"A{row_number}"].value),
                summary_name=summary_name,
                amount=_decimal(
                    _cell_value(formula_sheet, value_sheet, f"C{row_number}"),
                    field=f"{sheet_name}!C{row_number}",
                ),
                notes=_text(value_sheet[f"D{row_number}"].value),
                source_sheet=sheet_name,
                source_row=row_number,
                source_cells=_source_cells(
                    formula_sheet, value_sheet, row_number, "ABCD"
                ),
                row_evidence=_row_evidence(evidence_by_row, sheet_name, row_number),
                header_evidence=_row_evidence(evidence_by_row, sheet_name, 3),
            )
        )

    cover_name = "1-封皮"
    cover_formula = formula_workbook[cover_name]
    cover_values = value_workbook[cover_name]
    cover_records = (
        (12, BoqSummaryRecord.Kind.TAX_RATE, "增值税税率", "D", True),
        (14, BoqSummaryRecord.Kind.TAX_AMOUNT, "增值税合价", "D", False),
        (18, BoqSummaryRecord.Kind.TAX_INCLUDED_TOTAL, "合同含增值税总价", "D", False),
    )
    for row_number, kind, summary_name, value_column, is_rate in cover_records:
        value = _decimal(
            _cell_value(cover_formula, cover_values, f"{value_column}{row_number}"),
            field=f"{cover_name}!{value_column}{row_number}",
        )
        records.append(
            BoqSummaryRecord(
                import_run=run,
                document_version=spec.document_version,
                project_name=spec.project_name.strip(),
                party_a_name=spec.party_a_name.strip(),
                party_a_group=spec.party_a_group.strip(),
                kind=kind,
                summary_name=summary_name,
                amount=None if is_rate else value,
                rate=value if is_rate else None,
                tax_included=kind == BoqSummaryRecord.Kind.TAX_INCLUDED_TOTAL,
                source_sheet=cover_name,
                source_row=row_number,
                source_cells=_source_cells(
                    cover_formula, cover_values, row_number, "BCD"
                ),
                row_evidence=_row_evidence(evidence_by_row, cover_name, row_number),
                header_evidence=_row_evidence(evidence_by_row, cover_name, 3),
            )
        )
    return records


def _validate_business_totals(item_records, summary_records) -> None:
    summaries_by_name = {record.summary_name: record for record in summary_records}
    required_names = {
        "开办费清单",
        "实体工程量清单",
        "东塔",
        "西塔",
        "照管费清单",
        "计日工工程量清单",
        "暂列清单",
        "不含增值税总计",
        "增值税税率",
        "增值税合价",
        "合同含增值税总价",
    }
    missing = sorted(required_names - set(summaries_by_name))
    if missing:
        raise ValueError(f"缺少必需汇总记录：{missing}")

    item_sums = Counter()
    for record in item_records:
        if record.kind == BoqItemRecord.Kind.ENTITY:
            item_sums[record.tower_name] += _money(record.total_price)
        elif record.kind == BoqItemRecord.Kind.OPENING:
            item_sums["开办费清单"] += _money(record.total_price)
        elif record.kind == BoqItemRecord.Kind.DAYWORK:
            item_sums["计日工工程量清单"] += _money(record.total_price)

    for name in ("东塔", "西塔", "开办费清单", "计日工工程量清单"):
        summary_amount = summaries_by_name[name].amount
        if summary_amount is None:
            raise ValueError(f"汇总记录 {name!r} 金额为空")
        _require_money_equal(
            summary_amount, item_sums[name], label=f"{name} 汇总"
        )

    entity_total = summaries_by_name["实体工程量清单"].amount
    if entity_total is None:
        raise ValueError("实体工程量清单汇总金额为空")
    _require_money_equal(
        entity_total,
        _money(summaries_by_name["东塔"].amount)
        + _money(summaries_by_name["西塔"].amount),
        label="实体工程量清单汇总",
    )

    pre_tax_total = summaries_by_name["不含增值税总计"].amount
    if pre_tax_total is None:
        raise ValueError("不含增值税总计金额为空")
    component_total = sum(
        (
            _money(summaries_by_name["开办费清单"].amount),
            _money(entity_total),
            _money(summaries_by_name["照管费清单"].amount),
            _money(summaries_by_name["计日工工程量清单"].amount),
            _money(summaries_by_name["暂列清单"].amount),
        ),
        Decimal("0.00"),
    )
    _require_money_equal(pre_tax_total, component_total, label="不含增值税总计")

    tax_rate = summaries_by_name["增值税税率"].rate
    tax_amount = summaries_by_name["增值税合价"].amount
    tax_included_total = summaries_by_name["合同含增值税总价"].amount
    if tax_rate is None or tax_amount is None or tax_included_total is None:
        raise ValueError("封皮税率、税额或含税总价为空")
    calculated_tax = (pre_tax_total * tax_rate).quantize(
        Decimal("0.01"), rounding=ROUND_HALF_UP
    )
    _require_money_equal(tax_amount, calculated_tax, label="增值税合价")
    _require_money_equal(
        tax_included_total,
        _money(pre_tax_total) + _money(tax_amount),
        label="合同含增值税总价",
    )


def _sheet_snapshots(
    run,
    formula_workbook,
    value_workbook,
    item_records,
    summary_records,
) -> list[BoqSheetSnapshot]:
    imported_counts = Counter(record.source_sheet for record in item_records)
    imported_counts.update(record.analysis_source_sheet for record in item_records if record.analysis_source_sheet)
    imported_counts.update(record.source_sheet for record in summary_records)
    empty_template_kinds = {
        BoqSheetSnapshot.Kind.OPENING_APPENDIX,
        BoqSheetSnapshot.Kind.MANAGEMENT,
        BoqSheetSnapshot.Kind.PROVISIONAL,
    }
    snapshots = []
    for sheet_name, (kind, header_row) in _SHEET_PROFILES.items():
        formula_sheet = formula_workbook[sheet_name]
        value_sheet = value_workbook[sheet_name]
        nonempty_rows = 0
        formula_count = 0
        for row_number in range(1, formula_sheet.max_row + 1):
            row_has_value = False
            for column_number in range(1, formula_sheet.max_column + 1):
                coordinate = f"{get_column_letter(column_number)}{row_number}"
                raw_value = formula_sheet[coordinate].value
                cached_value = value_sheet[coordinate].value
                if raw_value is not None or cached_value is not None:
                    row_has_value = True
                formula_count += int(
                    isinstance(raw_value, str) and raw_value.startswith("=")
                )
            nonempty_rows += int(row_has_value)
        imported_record_count = imported_counts[sheet_name]
        snapshots.append(
            BoqSheetSnapshot(
                import_run=run,
                source_sheet=sheet_name,
                kind=kind,
                header_row=header_row,
                nonempty_row_count=nonempty_rows,
                formula_count=formula_count,
                imported_record_count=imported_record_count,
                is_empty_template=(
                    kind in empty_template_kinds and imported_record_count == 0
                ),
                metadata={
                    "max_row": formula_sheet.max_row,
                    "max_column": formula_sheet.max_column,
                    "hidden_columns": [
                        column
                        for column, dimension in formula_sheet.column_dimensions.items()
                        if dimension.hidden
                    ],
                    "hidden_rows": [
                        row_number
                        for row_number, dimension in formula_sheet.row_dimensions.items()
                        if dimension.hidden
                    ],
                },
            )
        )
    return snapshots


def _persist_boq_import(
    run: BoqImportRun,
    snapshots: list[BoqSheetSnapshot],
    item_records: list[BoqItemRecord],
    summary_records: list[BoqSummaryRecord],
    warnings: list[str],
) -> BoqImportRun:
    with transaction.atomic():
        BoqSheetSnapshot.objects.bulk_create(snapshots, batch_size=100)
        BoqItemRecord.objects.bulk_create(item_records, batch_size=500)
        BoqSummaryRecord.objects.bulk_create(summary_records, batch_size=100)
        BoqImportRun.objects.filter(pk=run.pk).update(
            status=BoqImportRun.Status.SUCCEEDED,
            imported_sheet_count=len(snapshots),
            imported_item_count=len(item_records),
            imported_summary_count=len(summary_records),
            warnings=warnings,
            finished_at=timezone.now(),
        )
    run.refresh_from_db()
    return run


def import_boq_spec(import_spec_id) -> BoqImportRun:
    spec = BoqImportSpec.objects.select_related(
        "document_version__document", "source_file"
    ).get(pk=import_spec_id)
    spec.full_clean()
    run = BoqImportRun.objects.create(
        import_spec=spec, status=BoqImportRun.Status.RUNNING
    )
    try:
        parser_run = (
            spec.source_file.parser_runs.filter(status=ParserRun.Status.SUCCEEDED)
            .order_by("-finished_at")
            .first()
        )
        if parser_run is None:
            raise ValueError("工程量清单源文件尚无成功 ParserRun")
        if spec.profile == BoqImportSpec.Profile.CRLAND_LIGHTING_XLS_V1:
            from .boq_xls import build_crland_lighting_xls_import

            snapshots, item_records, summary_records, warnings = (
                build_crland_lighting_xls_import(run, spec, parser_run)
            )
            return _persist_boq_import(
                run,
                snapshots,
                item_records,
                summary_records,
                warnings,
            )
        if spec.profile != BoqImportSpec.Profile.CRLAND_GENERAL_V1:
            raise ValueError(f"不受支持的工程量清单 profile：{spec.profile!r}")
        formula_workbook = load_workbook(
            Path(spec.source_file.file.path),
            data_only=False,
            read_only=False,
            keep_links=False,
        )
        value_workbook = load_workbook(
            Path(spec.source_file.file.path),
            data_only=True,
            read_only=False,
            keep_links=False,
        )
        _validate_workbook_profile(formula_workbook, value_workbook)

        cover_project_name = _text(value_workbook["1-封皮"]["D6"].value)
        if cover_project_name != spec.project_name.strip():
            raise ValueError(
                f"导入配置项目名称 {spec.project_name!r} 与 1-封皮!D6 "
                f"{cover_project_name!r} 不一致"
            )
        cover_party_a_name = _text(value_workbook["1-封皮"]["D5"].value)
        if cover_party_a_name and cover_party_a_name != spec.party_a_name.strip():
            raise ValueError(
                f"导入配置甲方名称 {spec.party_a_name!r} 与 1-封皮!D5 "
                f"{cover_party_a_name!r} 不一致"
            )

        evidence_by_row = _evidence_index(parser_run)
        item_records = _entity_records(
            run, spec, formula_workbook, value_workbook, evidence_by_row
        )
        item_records.extend(
            _opening_records(
                run, spec, formula_workbook, value_workbook, evidence_by_row
            )
        )
        item_records.extend(
            _opening_appendix_records(
                run, spec, formula_workbook, value_workbook, evidence_by_row
            )
        )
        item_records.extend(
            _daywork_records(
                run, spec, formula_workbook, value_workbook, evidence_by_row
            )
        )
        summary_records = _summary_records(
            run, spec, formula_workbook, value_workbook, evidence_by_row
        )
        _validate_business_totals(item_records, summary_records)
        snapshots = _sheet_snapshots(
            run,
            formula_workbook,
            value_workbook,
            item_records,
            summary_records,
        )
        warnings = _manifest_warnings(value_workbook)
        warnings.append(
            "公式值来自 XLSX 保存缓存，尚未受控重算；本次记录只进入 Staging"
        )
        if not cover_party_a_name:
            warnings.append(
                "1-封皮!D5 未填写建设单位；甲方名称来自人工确认的 BoqImportSpec"
            )

        return _persist_boq_import(
            run,
            snapshots,
            item_records,
            summary_records,
            warnings,
        )
    except Exception as exc:
        BoqImportRun.objects.filter(pk=run.pk).update(
            status=BoqImportRun.Status.FAILED,
            error=f"{type(exc).__name__}: {exc}"[:8_000],
            finished_at=timezone.now(),
        )
        run.refresh_from_db()
        return run


def trust_boq_item_records(queryset, *, reviewer) -> int:
    return queryset.filter(
        status=BoqItemRecord.Status.STAGING,
        import_run__status=BoqImportRun.Status.SUCCEEDED,
    ).update(
        status=BoqItemRecord.Status.TRUSTED,
        reviewed_by=reviewer,
        reviewed_at=timezone.now(),
    )


def trust_boq_summary_records(queryset, *, reviewer) -> int:
    return queryset.filter(
        status=BoqSummaryRecord.Status.STAGING,
        import_run__status=BoqImportRun.Status.SUCCEEDED,
    ).update(
        status=BoqSummaryRecord.Status.TRUSTED,
        reviewed_by=reviewer,
        reviewed_at=timezone.now(),
    )
