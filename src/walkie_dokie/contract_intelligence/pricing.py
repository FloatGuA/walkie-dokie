"""Declarative XLSX price import, exact query and deterministic calculation."""

from __future__ import annotations

import datetime as dt
import re
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.exceptions import ValidationError
from django.db import transaction
from django.db.models import Q
from django.utils import timezone
from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string

from .models import (
    EvidenceUnit,
    IndexBuildDocument,
    KnowledgeProject,
    PriceImportRun,
    PriceMappingSpec,
    PriceRecord,
)

MAPPABLE_PRICE_FIELDS = frozenset(
    {
        "product_code",
        "product_name",
        "region",
        "customer",
        "channel",
        "price_kind",
        "unit_price",
        "currency",
        "unit",
        "tax_included",
        "valid_from",
        "valid_to",
        "minimum_quantity",
        "maximum_quantity",
        "notes",
    }
)
_COLUMN_RE = re.compile(r"^[A-Z]{1,3}$")


def validate_mapping_spec(field_columns: dict) -> None:
    if not isinstance(field_columns, dict):
        raise ValidationError("field_columns 必须是 JSON object")
    unknown = set(field_columns) - MAPPABLE_PRICE_FIELDS
    if unknown:
        raise ValidationError(f"MappingSpec 包含不支持的字段：{sorted(unknown)}")
    missing = {"product_name", "unit_price"} - set(field_columns)
    if missing:
        raise ValidationError(f"MappingSpec 缺少必填字段：{sorted(missing)}")
    columns = []
    for field, column in field_columns.items():
        if not isinstance(column, str) or not _COLUMN_RE.fullmatch(column.upper()):
            raise ValidationError(f"{field} 必须映射到 Excel 列字母")
        columns.append(column.upper())
    if len(set(columns)) != len(columns):
        raise ValidationError("一个 Excel 列不能映射到多个价格字段")


def _decimal(value, *, field: str) -> Decimal | None:
    if value in {None, ""}:
        return None
    cleaned = str(value).strip().replace(",", "").replace("￥", "").replace("¥", "")
    try:
        return Decimal(cleaned)
    except InvalidOperation as exc:
        raise ValueError(f"{field} 不是合法十进制数：{value!r}") from exc


def _date(value, *, field: str) -> dt.date | None:
    if value in {None, ""}:
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    cleaned = str(value).strip().replace("年", "-").replace("月", "-").replace("日", "")
    cleaned = cleaned.replace("/", "-").replace(".", "-")
    try:
        return dt.date.fromisoformat(cleaned)
    except ValueError as exc:
        raise ValueError(f"{field} 不是可识别日期：{value!r}") from exc


def _boolean(value, *, field: str) -> bool | None:
    if value in {None, ""}:
        return None
    normalized = str(value).strip().casefold()
    if normalized in {"是", "含税", "true", "yes", "y", "1"}:
        return True
    if normalized in {"否", "不含税", "未税", "false", "no", "n", "0"}:
        return False
    raise ValueError(f"{field} 不是可识别布尔值：{value!r}")


def _text(value) -> str:
    return "" if value is None else str(value).strip()


def _row_evidence(run_ids, sheet_name: str, row_number: int) -> EvidenceUnit:
    candidates = list(
        EvidenceUnit.objects.filter(
            parser_run_id__in=run_ids,
            kind="spreadsheet_row",
            source_anchor__sheet=sheet_name,
            source_anchor__row=row_number,
        ).order_by("-created_at")[:2]
    )
    if len(candidates) != 1:
        raise ValueError(
            f"工作表 {sheet_name!r} 第 {row_number} 行没有唯一 Evidence Unit，必须先解析并选择单一 ParserRun"
        )
    return candidates[0]


def import_price_mapping(mapping_spec_id) -> PriceImportRun:
    spec = PriceMappingSpec.objects.select_related(
        "document_version__document", "source_file"
    ).get(pk=mapping_spec_id)
    spec.full_clean()
    run = PriceImportRun.objects.create(
        mapping_spec=spec, status=PriceImportRun.Status.RUNNING
    )
    try:
        parser_runs = list(
            spec.source_file.parser_runs.filter(status="succeeded").order_by("-finished_at")[:1]
        )
        if not parser_runs:
            raise ValueError("价目表源文件尚无成功 ParserRun")
        parser_run_ids = [item.id for item in parser_runs]
        workbook = load_workbook(
            Path(spec.source_file.file.path),
            read_only=False,
            data_only=True,
            keep_links=False,
        )
        if spec.sheet_name not in workbook.sheetnames:
            raise ValueError(f"XLSX 不存在工作表：{spec.sheet_name!r}")
        sheet = workbook[spec.sheet_name]
        if sheet.sheet_state != "visible":
            raise ValueError("MappingSpec 不允许导入隐藏工作表")
        end_row = spec.data_end_row or sheet.max_row
        field_columns = {
            field: column.upper() for field, column in spec.field_columns.items()
        }
        header_evidence = _row_evidence(
            parser_run_ids, spec.sheet_name, spec.header_row
        )
        records: list[PriceRecord] = []
        rejected: list[dict] = []
        for row_number in range(spec.data_start_row, end_row + 1):
            values = {
                field: sheet[f"{column}{row_number}"].value
                for field, column in field_columns.items()
            }
            if all(value in {None, ""} for value in values.values()):
                continue
            try:
                product_name = _text(values.get("product_name"))
                unit_price = _decimal(values.get("unit_price"), field="unit_price")
                if not product_name:
                    raise ValueError("product_name 为空")
                if unit_price is None:
                    raise ValueError("unit_price 为空")
                row_evidence = _row_evidence(
                    parser_run_ids, spec.sheet_name, row_number
                )
                records.append(
                    PriceRecord(
                        import_run=run,
                        document_version=spec.document_version,
                        product_code=_text(values.get("product_code")),
                        product_name=product_name,
                        region=_text(values.get("region")),
                        customer=_text(values.get("customer")),
                        channel=_text(values.get("channel")),
                        price_kind=_text(values.get("price_kind")),
                        unit_price=unit_price,
                        currency=_text(values.get("currency")),
                        unit=_text(values.get("unit")),
                        tax_included=_boolean(values.get("tax_included"), field="tax_included"),
                        valid_from=_date(values.get("valid_from"), field="valid_from"),
                        valid_to=_date(values.get("valid_to"), field="valid_to"),
                        minimum_quantity=_decimal(
                            values.get("minimum_quantity"), field="minimum_quantity"
                        ),
                        maximum_quantity=_decimal(
                            values.get("maximum_quantity"), field="maximum_quantity"
                        ),
                        notes=_text(values.get("notes")),
                        source_sheet=spec.sheet_name,
                        source_row=row_number,
                        source_cells={
                            field: {
                                "coordinate": f"{column}{row_number}",
                                "raw_value": (
                                    None if values[field] is None else str(values[field])
                                ),
                            }
                            for field, column in field_columns.items()
                        },
                        row_evidence=row_evidence,
                        header_evidence=header_evidence,
                    )
                )
            except Exception as exc:
                rejected.append({"row": row_number, "error": str(exc)[:1_000]})
        if rejected:
            raise ValidationError(
                f"有 {len(rejected)} 行不能确定性导入；修正 MappingSpec 或原文件后重试"
            )
        with transaction.atomic():
            PriceRecord.objects.bulk_create(records, batch_size=500)
            PriceImportRun.objects.filter(pk=run.pk).update(
                status=PriceImportRun.Status.SUCCEEDED,
                imported_count=len(records),
                rejected_rows=[],
                finished_at=timezone.now(),
            )
        run.refresh_from_db()
        return run
    except Exception as exc:
        PriceImportRun.objects.filter(pk=run.pk).update(
            status=PriceImportRun.Status.FAILED,
            rejected_rows=locals().get("rejected", []),
            error=f"{type(exc).__name__}: {exc}"[:8_000],
            finished_at=timezone.now(),
        )
        run.refresh_from_db()
        return run


def trust_price_records(queryset, *, reviewer) -> int:
    now = timezone.now()
    return queryset.filter(
        status=PriceRecord.Status.STAGING,
        import_run__status=PriceImportRun.Status.SUCCEEDED,
    ).update(
        status=PriceRecord.Status.TRUSTED,
        reviewed_by=reviewer,
        reviewed_at=now,
    )


@dataclass(frozen=True, slots=True)
class PriceQuery:
    product_name: str | None = None
    product_code: str | None = None
    region: str | None = None
    customer: str | None = None
    channel: str | None = None
    price_kind: str | None = None
    on_date: dt.date | None = None
    quantity: Decimal | None = None

    def __post_init__(self) -> None:
        if not (self.product_name or self.product_code):
            raise ValueError("PriceQuery 必须提供 product_name 或 product_code")
        if self.quantity is not None and self.quantity <= 0:
            raise ValueError("quantity 必须大于 0")


@dataclass(frozen=True, slots=True)
class PriceLookup:
    status: str
    records: tuple[PriceRecord, ...]
    reason: str
    missing_dimension: str | None = None


def query_published_prices(project_id, query: PriceQuery) -> PriceLookup:
    project = KnowledgeProject.objects.select_related("current_index_build").get(
        pk=project_id, is_active=True
    )
    build = project.current_index_build
    if build is None or build.status != build.Status.PUBLISHED:
        raise LookupError("项目没有已发布 IndexBuild")
    version_ids = IndexBuildDocument.objects.filter(index_build=build).values_list(
        "document_version_id", flat=True
    )
    records = PriceRecord.objects.filter(
        document_version_id__in=version_ids,
        status=PriceRecord.Status.TRUSTED,
    ).select_related("row_evidence", "header_evidence", "document_version__document")
    if query.product_code:
        records = records.filter(product_code=query.product_code)
    if query.product_name:
        records = records.filter(product_name=query.product_name)
    for field in ("region", "customer", "channel", "price_kind"):
        value = getattr(query, field)
        if value:
            records = records.filter(**{field: value})
    if query.on_date:
        records = records.filter(
            Q(valid_from__isnull=True) | Q(valid_from__lte=query.on_date),
            Q(valid_to__isnull=True) | Q(valid_to__gte=query.on_date),
        )
    if query.quantity is not None:
        records = records.filter(
            Q(minimum_quantity__isnull=True) | Q(minimum_quantity__lte=query.quantity),
            Q(maximum_quantity__isnull=True) | Q(maximum_quantity__gte=query.quantity),
        )
    matches = tuple(records[:101])
    if not matches:
        return PriceLookup("not_found", (), "没有满足全部条件的可信价格记录")
    if len(matches) > 100:
        return PriceLookup("ambiguous", (), "匹配记录过多，需要补充查询条件")

    for field in ("region", "customer", "channel", "price_kind"):
        if getattr(query, field):
            continue
        values = {getattr(record, field) for record in matches if getattr(record, field)}
        if len(values) > 1:
            return PriceLookup(
                "ambiguous", matches, f"不同{field}对应不同价格", field
            )
    price_facts = {
        (
            record.unit_price,
            record.currency,
            record.unit,
            record.tax_included,
            record.valid_from,
            record.valid_to,
        )
        for record in matches
    }
    if len(price_facts) > 1:
        return PriceLookup("conflict", matches, "相同查询条件存在多个价格事实")
    return PriceLookup("found", matches, "唯一价格事实")


def calculate_total(unit_price: Decimal, quantity: Decimal) -> dict:
    if quantity <= 0:
        raise ValueError("quantity 必须大于 0")
    total = unit_price * quantity
    return {
        "formula": "unit_price * quantity",
        "operands": {
            "unit_price": str(unit_price),
            "quantity": str(quantity),
        },
        "result": str(total),
        "rounding": "none",
    }
